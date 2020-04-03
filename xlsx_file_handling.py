from openpyxl import load_workbook, Workbook
import itertools

def read_xlsx(file_path, start_row=None, sheet_name=None):
    '''Will return the values of excel file as list of lists'''
    wb = load_workbook(file_path)
    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb.active
    rows_iter = ws.iter_rows(min_row=start_row)
    all_values = [[cell.value for cell in list(row)] for row in rows_iter]
    return all_values


def read_xlsx_list_of_dict(file_path, start_row=None, sheet_name=None):
    '''Will return the values of excel file as list of dictionaries'''
    data = read_xlsx(file_path, start_row=start_row,sheet_name=sheet_name)
    output_data = [{data[0][n]:cell for n,
                    cell in enumerate(row)}for row in data[1:]]
    return output_data


def write_xlsx(file_name, data_to_store):
    '''Will write list of lists into a XLSX file'''
    wb = Workbook()
    ws = wb.active
    for row in data_to_store:
        ws.append(row)
    wb.save(filename=file_name)


def write_list_of_dict_xlsx(filename, data_to_store):
    '''Will write list of dictionaries into a XLSX file'''
    wb = Workbook(write_only=True)
    ws = wb.create_sheet()

    headers = list(dict.fromkeys(itertools.chain.from_iterable(data_to_store)))
    ws.append(headers)

    for elements in data_to_store:
        ws.append([elements.get(h) for h in headers])

    wb.save(filename)
